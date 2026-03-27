#!/usr/bin/env python3
"""
SA10M Contest Manager - Tkinter Desktop UI

Simple GUI to manage contest logs: import, cross-check, and scoring.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import sys
import io
import os
from datetime import datetime
from pathlib import Path


# ─────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────

DEFAULT_DB = "sa10_contest.db"


def _utc_stamp() -> str:
    return datetime.utcnow().strftime("%H:%M:%S")


class TextRedirector(io.TextIOBase):
    """Redirect stdout/stderr to a ScrolledText widget."""

    def __init__(self, widget: scrolledtext.ScrolledText, tag: str = "normal"):
        self.widget = widget
        self.tag = tag

    def write(self, text: str) -> int:
        self.widget.configure(state="normal")
        self.widget.insert("end", text, self.tag)
        self.widget.see("end")
        self.widget.configure(state="disabled")
        return len(text)

    def flush(self):
        pass


# ─────────────────────────────────────────────────────────────────────
#  Callsign area classifier  (ITU prefix tables)
# ─────────────────────────────────────────────────────────────────────

import re as _re

# Argentine allocations: LU, LW, LO, LP, LQ (2-letter), AY, AZ
# and L2-L9 (1 letter + digit)
_ARGENTINA_PREFIXES = frozenset({
    'LU', 'LW', 'LO', 'LP', 'LQ',
    'AY', 'AZ',
})

# South-American allocations (everything except Argentina)
_SA_PREFIXES = frozenset({
    # Brazil
    'PP', 'PR', 'PS', 'PT', 'PU', 'PY',
    'ZV', 'ZW', 'ZX', 'ZY', 'ZZ',
    # Chile
    'CA', 'CB', 'CC', 'CD', 'CE', 'XQ',
    # Venezuela (letter-prefix allocs)
    'YV', 'YW', 'YX',
    # Colombia
    'HJ', 'HK',
    # Peru
    'OA', 'OB', 'OC',
    # Ecuador
    'HC', 'HD',
    # Bolivia
    'CP',
    # Paraguay
    'ZP',
    # Uruguay
    'CV', 'CX',
    # Suriname
    'PZ',
    # French Guiana
    'FY',
    # Aruba
    'P4',
})

# 2-char prefixes that **start with a digit** (digit-prefix countries in SA)
_SA_DIGIT_PREFIXES = frozenset({'4M', '8R', '9Y', '9Z'})


def _callsign_area(callsign: str) -> str:
    """
    Return 'Argentina', 'South America', or 'DX' for a given callsign.
    Handles /P, /QRP, /MM suffixes and prefix/call formats.
    """
    cs = callsign.strip().upper()
    if '/' in cs:
        cs = max(cs.split('/'), key=len)   # keep the longest segment
    if not cs:
        return 'DX'

    # --- prefixes that start with a digit (4M=Venezuela, 8R=Guyana…)
    if cs[0].isdigit():
        itu2 = cs[:2]
        return 'South America' if itu2 in _SA_DIGIT_PREFIXES else 'DX'

    # --- normal letter-prefix: extract letters before first digit
    m = _re.match(r'^([A-Z]+)', cs)
    if not m:
        return 'DX'
    letters = m.group(1)

    # L + digit 2-9  →  Argentine special allocations (L2AA…L9ZZ)
    if letters == 'L':
        nxt = cs[1:2]
        return 'Argentina' if (nxt.isdigit() and nxt >= '2') else 'DX'

    if letters in _ARGENTINA_PREFIXES:
        return 'Argentina'
    if letters in _SA_PREFIXES:
        return 'South America'
    return 'DX'


# ─────────────────────────────────────────────────────────────────────
#  Main Application
# ─────────────────────────────────────────────────────────────────────

class SA10App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SA10M Contest Manager")
        self.geometry("960x720")
        self.resizable(True, True)
        self.configure(bg="#f0f0f0")

        # State
        self._db_path = tk.StringVar(value=DEFAULT_DB)
        self._selected_contest_id = tk.IntVar(value=0)
        self._selected_contest_name = tk.StringVar(value="(none)")
        self._running = False

        self._build_menu()
        self._build_header()
        self._build_notebook()
        self._build_log_pane()
        self._build_statusbar()

        # Redirect stdout/print to the log pane
        sys.stdout = TextRedirector(self._log_text, "stdout")

        # Ensure the database exists and has all tables before first use
        self._init_database()

        # Load contests on startup
        self.after(200, self._refresh_contests)

    # ── Database bootstrap ────────────────────────────────────────────

    def _init_database(self):
        """Create the database file and all tables if they don't exist yet."""
        try:
            from src.database.db_manager import DatabaseManager
            db = DatabaseManager(self._db_path.get())
            db.create_all_tables()
        except Exception as e:
            # Non-fatal: user can still choose a different DB via File menu
            self._log(f"Warning: could not initialise database '{self._db_path.get()}': {e}", "warn")

    # ── Menu ──────────────────────────────────────────────────────────

    def _build_menu(self):
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Select Database…", command=self._select_db)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.destroy)
        menubar.add_cascade(label="File", menu=file_menu)
        self.config(menu=menubar)

    # ── Header ────────────────────────────────────────────────────────

    def _build_header(self):
        hdr = tk.Frame(self, bg="#1a3a5c", padx=10, pady=6)
        hdr.pack(fill="x")

        tk.Label(hdr, text="SA10M Contest Manager", font=("Segoe UI", 14, "bold"),
                 bg="#1a3a5c", fg="white").pack(side="left")

        # DB path display
        right = tk.Frame(hdr, bg="#1a3a5c")
        right.pack(side="right")
        tk.Label(right, text="DB:", bg="#1a3a5c", fg="#aac8e0",
                 font=("Segoe UI", 9)).pack(side="left")
        tk.Label(right, textvariable=self._db_path, bg="#1a3a5c", fg="white",
                 font=("Segoe UI", 9)).pack(side="left", padx=(2, 8))
        tk.Button(right, text="Change", command=self._select_db,
                  bg="#2d6a9f", fg="white", relief="flat", padx=6,
                  font=("Segoe UI", 8)).pack(side="left")

    # ── Notebook tabs ─────────────────────────────────────────────────

    def _build_notebook(self):
        self._nb = ttk.Notebook(self)
        self._nb.pack(fill="both", expand=False, padx=8, pady=(6, 0))

        self._tab_contests = ttk.Frame(self._nb)
        self._tab_import   = ttk.Frame(self._nb)
        self._tab_xcheck   = ttk.Frame(self._nb)
        self._tab_scoring  = ttk.Frame(self._nb)
        self._tab_results  = ttk.Frame(self._nb)

        self._nb.add(self._tab_contests, text="  Contests  ")
        self._nb.add(self._tab_import,   text="  Import Logs  ")
        self._nb.add(self._tab_xcheck,   text="  Cross-Check  ")
        self._nb.add(self._tab_scoring,  text="  Scoring  ")
        self._nb.add(self._tab_results,  text="  Leaderboard  ")

        self._build_tab_contests()
        self._build_tab_import()
        self._build_tab_xcheck()
        self._build_tab_scoring()
        self._build_tab_results()

    # ── Log pane ──────────────────────────────────────────────────────

    def _build_log_pane(self):
        frame = tk.LabelFrame(self, text="Output Log", font=("Segoe UI", 9),
                              bg="#f0f0f0", padx=4, pady=4)
        frame.pack(fill="both", expand=True, padx=8, pady=(4, 2))

        self._log_text = scrolledtext.ScrolledText(
            frame, height=12, state="disabled", wrap="word",
            font=("Consolas", 9), bg="#1e1e1e", fg="#d4d4d4",
            insertbackground="white"
        )
        self._log_text.pack(fill="both", expand=True)
        self._log_text.tag_config("stdout", foreground="#d4d4d4")
        self._log_text.tag_config("warn",   foreground="#f0c040")
        self._log_text.tag_config("ok",     foreground="#6bc46b")
        self._log_text.tag_config("error",  foreground="#f07070")

        btn_row = tk.Frame(frame, bg="#f0f0f0")
        btn_row.pack(fill="x", pady=(2, 0))
        tk.Button(btn_row, text="Clear Log", command=self._clear_log,
                  font=("Segoe UI", 8), padx=6).pack(side="right")

    def _clear_log(self):
        self._log_text.configure(state="normal")
        self._log_text.delete("1.0", "end")
        self._log_text.configure(state="disabled")

    def _log(self, msg: str, tag: str = "stdout"):
        self._log_text.configure(state="normal")
        self._log_text.insert("end", f"[{_utc_stamp()}] {msg}\n", tag)
        self._log_text.see("end")
        self._log_text.configure(state="disabled")

    # ── Status bar ────────────────────────────────────────────────────

    def _build_statusbar(self):
        bar = tk.Frame(self, bg="#dce3ea", relief="sunken", bd=1)
        bar.pack(fill="x", side="bottom")
        self._status_var = tk.StringVar(value="Ready")
        tk.Label(bar, textvariable=self._status_var, anchor="w",
                 bg="#dce3ea", font=("Segoe UI", 8), padx=6).pack(side="left")
        self._progress = ttk.Progressbar(bar, mode="indeterminate", length=140)
        self._progress.pack(side="right", padx=6, pady=2)

    def _set_status(self, msg: str, busy: bool = False):
        self._status_var.set(msg)
        if busy:
            self._progress.start(12)
        else:
            self._progress.stop()

    # ─────────────────────────────────────────────────────────────────
    #  TAB: Contests
    # ─────────────────────────────────────────────────────────────────

    def _build_tab_contests(self):
        f = self._tab_contests

        # ── Create contest form ──
        create_frame = tk.LabelFrame(f, text="Create New Contest",
                                     font=("Segoe UI", 9), padx=8, pady=6)
        create_frame.pack(fill="x", padx=10, pady=(10, 4))

        labels  = ["Name:", "Slug:", "Start (YYYY-MM-DD HH:MM):", "End (YYYY-MM-DD HH:MM):"]
        defaults = ["SA10M 2025", "sa10m-2025", "2025-03-08 00:00", "2025-03-09 23:59"]
        self._contest_entries = {}
        for col, (lbl, dflt) in enumerate(zip(labels, defaults)):
            tk.Label(create_frame, text=lbl, font=("Segoe UI", 9)).grid(
                row=0, column=col*2, sticky="e", padx=(8, 2), pady=4)
            var = tk.StringVar(value=dflt)
            ent = tk.Entry(create_frame, textvariable=var, width=22,
                           font=("Segoe UI", 9))
            ent.grid(row=0, column=col*2+1, sticky="w", padx=(0, 8))
            self._contest_entries[lbl] = var

        tk.Button(create_frame, text="Create Contest", command=self._create_contest,
                  bg="#2d6a9f", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4).grid(
            row=1, column=0, columnspan=8, pady=(6, 2))

        # ── Contest list ──
        list_frame = tk.LabelFrame(f, text="Existing Contests",
                                   font=("Segoe UI", 9), padx=8, pady=6)
        list_frame.pack(fill="both", expand=True, padx=10, pady=(4, 10))

        btn_row = tk.Frame(list_frame)
        btn_row.pack(fill="x", pady=(0, 4))
        tk.Button(btn_row, text="↻ Refresh", command=self._refresh_contests,
                  font=("Segoe UI", 8), padx=6).pack(side="left")
        tk.Button(btn_row, text="Select Active", command=self._select_active_contest,
                  font=("Segoe UI", 8), bg="#2d6a9f", fg="white",
                  relief="flat", padx=6).pack(side="left", padx=4)
        tk.Button(btn_row, text="Delete", command=self._delete_contest,
                  font=("Segoe UI", 8), bg="#9f2d2d", fg="white",
                  relief="flat", padx=6).pack(side="left")

        tk.Label(btn_row, text="Active contest:", font=("Segoe UI", 9)).pack(
            side="right", padx=(0, 4))
        tk.Label(btn_row, textvariable=self._selected_contest_name,
                 font=("Segoe UI", 9, "bold"), fg="#1a3a5c").pack(side="right")

        cols = ("ID", "Name", "Slug", "Start", "End", "Logs")
        self._contest_tree = ttk.Treeview(list_frame, columns=cols,
                                          show="headings", height=6)
        widths = (40, 200, 140, 140, 140, 60)
        for col, w in zip(cols, widths):
            self._contest_tree.heading(col, text=col)
            self._contest_tree.column(col, width=w, anchor="center" if col in ("ID","Logs") else "w")
        self._contest_tree.pack(fill="both", expand=True)

    def _refresh_contests(self):
        try:
            from src.database.db_manager import DatabaseManager
            from src.database.models import Contest
            from sqlalchemy import text as sql_text
            db = DatabaseManager(self._db_path.get())
            rows = []
            with db.get_session() as session:
                contests = session.query(Contest).order_by(Contest.id).all()
                for c in contests:
                    cnt = session.execute(
                        sql_text("SELECT COUNT(*) FROM logs WHERE contest_id=:cid"),
                        {"cid": c.id}
                    ).scalar() or 0
                    rows.append((
                        c.id, c.name, c.slug,
                        c.start_date.strftime("%Y-%m-%d %H:%M") if c.start_date else "",
                        c.end_date.strftime("%Y-%m-%d %H:%M")   if c.end_date   else "",
                        cnt
                    ))

            for item in self._contest_tree.get_children():
                self._contest_tree.delete(item)
            for row in rows:
                self._contest_tree.insert("", "end", values=row)
            self._log(f"Loaded {len(rows)} contest(s).", "ok")
        except Exception as e:
            self._log(f"Could not load contests: {e}", "error")

    def _select_active_contest(self):
        sel = self._contest_tree.selection()
        if not sel:
            messagebox.showwarning("Select a contest", "Please select a contest row first.")
            return
        row = self._contest_tree.item(sel[0])["values"]
        cid, cname = int(row[0]), str(row[1])
        self._selected_contest_id.set(cid)
        self._selected_contest_name.set(f"{cname} (ID={cid})")
        self._log(f"Active contest set to: {cname} (ID={cid})", "ok")
        self._set_status(f"Active contest: {cname}")

    def _create_contest(self):
        name  = self._contest_entries["Name:"].get().strip()
        slug  = self._contest_entries["Slug:"].get().strip()
        start = self._contest_entries["Start (YYYY-MM-DD HH:MM):"].get().strip()
        end   = self._contest_entries["End (YYYY-MM-DD HH:MM):"].get().strip()
        if not all([name, slug, start, end]):
            messagebox.showerror("Missing fields", "All fields are required.")
            return

        def _run():
            try:
                from src.database.db_manager import DatabaseManager
                from src.database.models import Contest
                db = DatabaseManager(self._db_path.get())
                with db.get_session() as session:
                    existing = session.query(Contest).filter(Contest.slug == slug).first()
                    if existing:
                        self._log(f"Contest with slug '{slug}' already exists.", "warn")
                        return
                    c = Contest(
                        name=name, slug=slug,
                        start_date=datetime.strptime(start, "%Y-%m-%d %H:%M"),
                        end_date=datetime.strptime(end,   "%Y-%m-%d %H:%M"),
                        rules_file="config/contests/sa10m.yaml"
                    )
                    session.add(c)
                    session.commit()
                    cid = c.id
                self._log(f"Contest '{name}' created (ID={cid}).", "ok")
                self._refresh_contests()
            except Exception as e:
                self._log(f"Error creating contest: {e}", "error")
            finally:
                self._set_status("Ready")

        self._set_status("Creating contest…", busy=True)
        threading.Thread(target=_run, daemon=True).start()

    def _delete_contest(self):
        sel = self._contest_tree.selection()
        if not sel:
            messagebox.showwarning("Select a contest", "Please select a contest row first.")
            return
        row  = self._contest_tree.item(sel[0])["values"]
        cid  = int(row[0])
        cname = str(row[1])
        if not messagebox.askyesno("Confirm Delete",
                                   f"Delete contest '{cname}' (ID={cid}) and ALL its logs/contacts?"):
            return

        def _run():
            try:
                from src.database.db_manager import DatabaseManager
                from sqlalchemy import text as sql_text
                db = DatabaseManager(self._db_path.get())
                with db.get_session() as session:
                    session.execute(sql_text("PRAGMA foreign_keys=OFF"))
                    session.execute(sql_text(
                        "DELETE FROM contacts WHERE log_id IN (SELECT id FROM logs WHERE contest_id=:cid)"),
                        {"cid": cid})
                    session.execute(sql_text(
                        "DELETE FROM scores WHERE log_id IN (SELECT id FROM logs WHERE contest_id=:cid)"),
                        {"cid": cid})
                    session.execute(sql_text("DELETE FROM logs WHERE contest_id=:cid"), {"cid": cid})
                    session.execute(sql_text("DELETE FROM contests WHERE id=:cid"),     {"cid": cid})
                    session.execute(sql_text("PRAGMA foreign_keys=ON"))
                    session.commit()
                self._log(f"Contest '{cname}' deleted.", "warn")
                self._refresh_contests()
            except Exception as e:
                self._log(f"Error deleting contest: {e}", "error")
            finally:
                self._set_status("Ready")

        self._set_status("Deleting…", busy=True)
        threading.Thread(target=_run, daemon=True).start()

    # ─────────────────────────────────────────────────────────────────
    #  TAB: Import Logs
    # ─────────────────────────────────────────────────────────────────

    def _build_tab_import(self):
        f = self._tab_import

        # Source selection
        src_frame = tk.LabelFrame(f, text="Log Source", font=("Segoe UI", 9),
                                  padx=8, pady=6)
        src_frame.pack(fill="x", padx=10, pady=(10, 4))

        self._import_mode = tk.StringVar(value="dir")
        tk.Radiobutton(src_frame, text="Directory (all .txt/.log files)",
                       variable=self._import_mode, value="dir",
                       font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=4)
        tk.Radiobutton(src_frame, text="Single file",
                       variable=self._import_mode, value="file",
                       font=("Segoe UI", 9)).grid(row=0, column=1, sticky="w", padx=4)

        self._import_path = tk.StringVar(value="logs_sa10m_2025")
        tk.Label(src_frame, text="Path:", font=("Segoe UI", 9)).grid(
            row=1, column=0, sticky="e", padx=(0, 4), pady=4)
        tk.Entry(src_frame, textvariable=self._import_path, width=60,
                 font=("Segoe UI", 9)).grid(row=1, column=1, sticky="w")
        tk.Button(src_frame, text="Browse…", command=self._browse_import,
                  font=("Segoe UI", 8), padx=6).grid(row=1, column=2, padx=4)

        # Options
        opt_frame = tk.LabelFrame(f, text="Options", font=("Segoe UI", 9),
                                  padx=8, pady=6)
        opt_frame.pack(fill="x", padx=10, pady=4)

        tk.Label(opt_frame, text="Contest ID:", font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky="e", padx=4)
        self._import_cid = tk.StringVar()
        self._import_cid.trace_add("write", lambda *_: None)

        def _sync_cid(*_):
            self._import_cid.set(str(self._selected_contest_id.get())
                                  if self._selected_contest_id.get() else "")
        self._selected_contest_id.trace_add("write", _sync_cid)

        tk.Entry(opt_frame, textvariable=self._import_cid, width=8,
                 font=("Segoe UI", 9)).grid(row=0, column=1, sticky="w", padx=4)
        tk.Label(opt_frame, text="(uses active contest if empty)",
                 font=("Segoe UI", 8), fg="gray").grid(row=0, column=2, sticky="w")

        self._import_clear_db = tk.BooleanVar(value=False)
        tk.Checkbutton(opt_frame, text="Clear ALL contest data before import (fresh start)",
                       variable=self._import_clear_db,
                       font=("Segoe UI", 9), fg="#9f2d2d").grid(
            row=1, column=0, columnspan=3, sticky="w", padx=4, pady=2)

        # Run button
        tk.Button(f, text="▶  Import Logs", command=self._run_import,
                  bg="#2d6a9f", fg="white", font=("Segoe UI", 11, "bold"),
                  relief="flat", padx=20, pady=8).pack(pady=10)

    def _browse_import(self):
        if self._import_mode.get() == "dir":
            path = filedialog.askdirectory(title="Select logs folder")
        else:
            path = filedialog.askopenfilename(
                title="Select log file",
                filetypes=[("Log files", "*.txt *.log *.cbr"), ("All", "*.*")])
        if path:
            self._import_path.set(path)

    def _run_import(self):
        if self._running:
            messagebox.showinfo("Busy", "An operation is already running.")
            return

        cid_str = self._import_cid.get().strip()
        if not cid_str and self._selected_contest_id.get():
            cid_str = str(self._selected_contest_id.get())
        if not cid_str:
            messagebox.showerror("No contest", "Set an active contest first (Contests tab).")
            return

        try:
            contest_id = int(cid_str)
        except ValueError:
            messagebox.showerror("Invalid", "Contest ID must be an integer.")
            return

        src_path   = self._import_path.get().strip()
        mode       = self._import_mode.get()
        clear_db   = self._import_clear_db.get()
        db_path    = self._db_path.get()

        def _run():
            self._running = True
            self._set_status("Importing logs…", busy=True)
            try:
                if clear_db:
                    self._log("Clearing contest data…", "warn")
                    from sqlalchemy import text as sql_text
                    from src.database.db_manager import DatabaseManager
                    db = DatabaseManager(db_path)
                    with db.get_session() as session:
                        session.execute(sql_text("PRAGMA foreign_keys=OFF"))
                        session.execute(sql_text(
                            "DELETE FROM contacts WHERE log_id IN "
                            "(SELECT id FROM logs WHERE contest_id=:cid)"), {"cid": contest_id})
                        session.execute(sql_text(
                            "DELETE FROM scores WHERE log_id IN "
                            "(SELECT id FROM logs WHERE contest_id=:cid)"), {"cid": contest_id})
                        session.execute(sql_text(
                            "DELETE FROM logs WHERE contest_id=:cid"), {"cid": contest_id})
                        session.execute(sql_text("PRAGMA foreign_keys=ON"))
                        session.commit()
                    self._log("Contest data cleared.", "warn")

                from src.database.db_manager import DatabaseManager
                from src.services.log_import_service import LogImportService
                db = DatabaseManager(db_path)

                if mode == "dir":
                    log_files = list(Path(src_path).glob("*.txt")) + \
                                list(Path(src_path).glob("*.log")) + \
                                list(Path(src_path).glob("*.cbr"))
                    self._log(f"Found {len(log_files)} log file(s) in '{src_path}'")
                    ok = err = 0
                    for lf in log_files:
                        with db.get_session() as session:
                            svc = LogImportService(db)
                            result = svc.import_cabrillo_file(str(lf), contest_id)
                        if result["success"]:
                            ok += 1
                        else:
                            err += 1
                            self._log(f"  SKIP {lf.name}: {result['message']}", "warn")
                    self._log(f"Import done — {ok} OK, {err} skipped/failed.", "ok")
                else:
                    with db.get_session() as session:
                        svc = LogImportService(db)
                        result = svc.import_cabrillo_file(src_path, contest_id)
                    if result["success"]:
                        self._log(
                            f"Imported {result['callsign']} — "
                            f"{result['qso_count']} QSOs.", "ok")
                    else:
                        self._log(f"Import failed: {result['message']}", "error")

                self._refresh_contests()
            except Exception as e:
                import traceback
                self._log(f"Import error: {e}", "error")
                self._log(traceback.format_exc(), "error")
            finally:
                self._running = False
                self._set_status("Ready")

        threading.Thread(target=_run, daemon=True).start()

    # ─────────────────────────────────────────────────────────────────
    #  TAB: Cross-Check
    # ─────────────────────────────────────────────────────────────────

    def _build_tab_xcheck(self):
        f = self._tab_xcheck

        info = tk.LabelFrame(f, text="Cross-Check Settings", font=("Segoe UI", 9),
                             padx=8, pady=8)
        info.pack(fill="x", padx=10, pady=(10, 4))

        tk.Label(info, text="Contest:", font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky="e", padx=4)
        tk.Label(info, textvariable=self._selected_contest_name,
                 font=("Segoe UI", 9, "bold"), fg="#1a3a5c").grid(
            row=0, column=1, sticky="w")

        self._xcheck_save_ubn = tk.BooleanVar(value=True)
        tk.Checkbutton(info, text="Save UBN reports to ubn_reports/ folder",
                       variable=self._xcheck_save_ubn,
                       font=("Segoe UI", 9)).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=4, pady=4)

        tk.Button(f, text="▶  Run Cross-Check", command=self._run_crosscheck,
                  bg="#2d8a46", fg="white", font=("Segoe UI", 11, "bold"),
                  relief="flat", padx=20, pady=8).pack(pady=10)

        tk.Label(f, text="Tip: Make sure logs are imported before running cross-check.",
                 font=("Segoe UI", 8), fg="gray").pack()

    def _run_crosscheck(self):
        if self._running:
            messagebox.showinfo("Busy", "An operation is already running.")
            return
        cid = self._selected_contest_id.get()
        if not cid:
            messagebox.showerror("No contest", "Select an active contest first.")
            return
        save_ubn = self._xcheck_save_ubn.get()
        db_path  = self._db_path.get()

        def _run():
            self._running = True
            self._set_status("Running cross-check…", busy=True)
            try:
                from src.database.db_manager import DatabaseManager
                from src.services.cross_check_service import CrossCheckService
                db = DatabaseManager(db_path)
                with db.get_session() as session:
                    svc = CrossCheckService(session)
                    t0 = datetime.now()
                    ubn_by_log = svc.check_all_logs(cid)
                    elapsed = (datetime.now() - t0).total_seconds()
                    self._log(
                        f"Cross-check complete in {elapsed:.1f}s — "
                        f"{len(ubn_by_log)} logs with issues.", "ok")

                    # Write NIL/busted results back to contacts table
                    svc.update_database_with_results(ubn_by_log)

                    if save_ubn:
                        out_dir = Path("ubn_reports")
                        out_dir.mkdir(exist_ok=True)
                        from src.services.ubn_report_generator import UBNReportGenerator
                        gen = UBNReportGenerator(session)
                        saved = 0
                        for log_id, entries in ubn_by_log.items():
                            try:
                                stats = svc.stats.get(log_id)
                                if stats is None:
                                    from src.services.cross_check_service import CrossCheckStats
                                    stats = CrossCheckStats(
                                        total_contacts=0, valid_contacts=0,
                                        unique_count=sum(1 for e in entries
                                                        if e.ubn_type.value == "unique"),
                                        busted_count=sum(1 for e in entries
                                                        if e.ubn_type.value == "busted"),
                                        nil_count=sum(1 for e in entries
                                                     if e.ubn_type.value == "nil"),
                                        matched_count=0)
                                callsign = entries[0].log_callsign if entries else f"log_{log_id}"
                                safe_callsign = callsign.replace("/", "_").replace("\\", "_")
                                report_txt = gen.generate_text_report(
                                    log_id, entries, stats, "SA10M 2025")
                                out_file = out_dir / f"{safe_callsign}_UBN.txt"
                                out_file.write_text(report_txt, encoding="utf-8")
                                saved += 1
                            except Exception as e2:
                                self._log(f"  Report error for log {log_id}: {e2}", "warn")
                        self._log(f"Saved {saved} UBN reports to ubn_reports/", "ok")

                    # Re-score affected logs so scores.invalid_qsos / not_in_log_qsos
                    # reflect the updated validation_status values from cross-check.
                    self._log("Re-scoring logs to update NIL/invalid counts…", "info")
                    from src.services.scoring_service import ScoringService
                    scoring_svc = ScoringService(session, 'sa10m')
                    scored, failed = 0, 0
                    for log_id in ubn_by_log:
                        try:
                            scoring_svc.score_log(log_id)
                            scored += 1
                        except Exception as se:
                            self._log(f"  Re-score error for log {log_id}: {se}", "warn")
                            failed += 1
                    self._log(
                        f"Re-scored {scored} logs" + (f" ({failed} failed)" if failed else ""),
                        "ok")

            except Exception as e:
                import traceback
                self._log(f"Cross-check error: {e}", "error")
                self._log(traceback.format_exc(), "error")
            finally:
                self._running = False
                self._set_status("Ready")

        threading.Thread(target=_run, daemon=True).start()

    # ─────────────────────────────────────────────────────────────────
    #  TAB: Scoring
    # ─────────────────────────────────────────────────────────────────

    def _build_tab_scoring(self):
        f = self._tab_scoring

        info = tk.LabelFrame(f, text="Scoring Settings", font=("Segoe UI", 9),
                             padx=8, pady=8)
        info.pack(fill="x", padx=10, pady=(10, 4))

        tk.Label(info, text="Contest:", font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky="e", padx=4)
        tk.Label(info, textvariable=self._selected_contest_name,
                 font=("Segoe UI", 9, "bold"), fg="#1a3a5c").grid(
            row=0, column=1, sticky="w")

        tk.Label(info, text="Rules slug:", font=("Segoe UI", 9)).grid(
            row=1, column=0, sticky="e", padx=4, pady=(4, 0))
        self._scoring_slug = tk.StringVar(value="sa10m")
        tk.Entry(info, textvariable=self._scoring_slug, width=20,
                 font=("Segoe UI", 9)).grid(row=1, column=1, sticky="w", pady=(4, 0))

        tk.Button(f, text="▶  Score All Logs", command=self._run_scoring,
                  bg="#7a3aaf", fg="white", font=("Segoe UI", 11, "bold"),
                  relief="flat", padx=20, pady=8).pack(pady=10)

        tk.Label(f,
                 text="Scoring will calculate QSO points, multipliers and final scores\n"
                      "for every log in the active contest.",
                 font=("Segoe UI", 9), fg="gray").pack()

    def _run_scoring(self):
        if self._running:
            messagebox.showinfo("Busy", "An operation is already running.")
            return
        cid = self._selected_contest_id.get()
        if not cid:
            messagebox.showerror("No contest", "Select an active contest first.")
            return
        slug    = self._scoring_slug.get().strip()
        db_path = self._db_path.get()

        def _run():
            self._running = True
            self._set_status("Scoring logs…", busy=True)
            ok = err = 0
            try:
                from src.database.db_manager import DatabaseManager
                from src.services.scoring_service import ScoringService
                from src.database.models import Log
                from sqlalchemy import select as sa_select
                db = DatabaseManager(db_path)
                with db.get_session() as session:
                    logs = session.execute(
                        sa_select(Log).where(Log.contest_id == cid)
                    ).scalars().all()
                    total = len(logs)
                    self._log(f"Scoring {total} logs…")
                    svc = ScoringService(session, slug)
                    for i, log in enumerate(logs, 1):
                        try:
                            result = svc.score_log(log.id)
                            ok += 1
                            if i % 50 == 0 or i == total:
                                self._log(
                                    f"  Progress: {i}/{total}  "
                                    f"({ok} OK, {err} errors)")
                        except Exception as e2:
                            err += 1
                            self._log(f"  Error on {log.callsign}: {e2}", "warn")
                    self._log(
                        f"Scoring complete — {ok} OK, {err} errors.", "ok")
                self._load_leaderboard(cid, db_path)
            except Exception as e:
                import traceback
                self._log(f"Scoring error: {e}", "error")
                self._log(traceback.format_exc(), "error")
            finally:
                self._running = False
                self._set_status("Ready")

        threading.Thread(target=_run, daemon=True).start()

    # ─────────────────────────────────────────────────────────────────
    #  TAB: Leaderboard
    # ─────────────────────────────────────────────────────────────────

    def _build_tab_results(self):
        f = self._tab_results

        btn_row = tk.Frame(f)
        btn_row.pack(fill="x", padx=10, pady=(8, 4))
        tk.Button(btn_row, text="↻ Refresh Leaderboard", command=self._refresh_leaderboard,
                  bg="#2d6a9f", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10).pack(side="left")
        tk.Button(btn_row, text="⬇ Export to Excel", command=self._export_excel,
                  bg="#217346", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10).pack(side="left", padx=(6, 0))
        tk.Button(btn_row, text="⬇ Export Scores CSV", command=self._export_scores_csv,
                  bg="#6a3d9a", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10).pack(side="left", padx=(6, 0))
        tk.Button(btn_row, text="⬇ QSO Report (Excel)", command=self._export_qso_excel,
                  bg="#b8560a", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10).pack(side="left", padx=(6, 0))

        # Callsign filter
        self._ldr_filter = tk.StringVar()
        tk.Label(btn_row, text="  Filter callsign:",
                 font=("Segoe UI", 9)).pack(side="left")
        filter_entry = tk.Entry(btn_row, textvariable=self._ldr_filter,
                                width=14, font=("Segoe UI", 9))
        filter_entry.pack(side="left", padx=4)
        filter_entry.bind("<KeyRelease>", lambda _: self._apply_leaderboard_filter())

        # Category filter
        tk.Label(btn_row, text="  Category:",
                 font=("Segoe UI", 9)).pack(side="left")
        self._ldr_cat_filter = tk.StringVar(value="All")
        self._ldr_cat_combo = ttk.Combobox(btn_row, textvariable=self._ldr_cat_filter,
                                           width=22, font=("Segoe UI", 9),
                                           state="readonly")
        self._ldr_cat_combo["values"] = ["All"]
        self._ldr_cat_combo.pack(side="left", padx=4)
        self._ldr_cat_combo.bind("<<ComboboxSelected>>", lambda _: self._apply_leaderboard_filter())

        # Area filter
        tk.Label(btn_row, text="  Area:",
                 font=("Segoe UI", 9)).pack(side="left")
        self._ldr_area_filter = tk.StringVar(value="World")
        area_combo = ttk.Combobox(btn_row, textvariable=self._ldr_area_filter,
                                  width=14, font=("Segoe UI", 9), state="readonly")
        area_combo["values"] = ["World", "Argentina", "South America", "DX"]
        area_combo.pack(side="left", padx=4)
        area_combo.bind("<<ComboboxSelected>>", lambda _: self._apply_leaderboard_filter())

        cols = ("#", "Callsign", "Category", "Final Score", "Total QSOs",
                "Valid QSOs", "Dupes", "Points", "Multipliers")
        self._ldr_tree = ttk.Treeview(f, columns=cols, show="headings", height=14)
        widths = (40, 120, 180, 100, 80, 80, 55, 80, 90)
        for col, w in zip(cols, widths):
            self._ldr_tree.heading(col, text=col,
                                   command=lambda c=col: self._sort_leaderboard(c))
            anchor = "w" if col in ("Callsign", "Category") else "center"
            self._ldr_tree.column(col, width=w, anchor=anchor)

        sb = ttk.Scrollbar(f, orient="vertical", command=self._ldr_tree.yview)
        self._ldr_tree.configure(yscrollcommand=sb.set)
        self._ldr_tree.pack(fill="both", expand=True, padx=10, side="left")
        sb.pack(side="right", fill="y", pady=0, padx=(0, 8))

        # internal cache
        self._ldr_data = []
        self._ldr_sort_col = "Category"
        self._ldr_sort_rev = False

    def _refresh_leaderboard(self):
        cid     = self._selected_contest_id.get()
        db_path = self._db_path.get()
        if not cid:
            messagebox.showinfo("No contest", "Select an active contest first.")
            return
        self._load_leaderboard(cid, db_path)

    def _load_leaderboard(self, cid: int, db_path: str):
        try:
            from src.database.db_manager import DatabaseManager
            from sqlalchemy import text as sql_text
            db = DatabaseManager(db_path)
            with db.get_session() as session:
                rows = session.execute(sql_text("""
                    SELECT l.callsign,
                           l.category_operator, l.category_mode, l.category_power,
                           s.final_score, s.total_qsos, s.valid_qsos,
                           s.duplicate_qsos, s.total_points, s.multipliers
                    FROM scores s
                    JOIN logs l ON s.log_id = l.id
                    WHERE l.contest_id = :cid
                    ORDER BY s.final_score DESC
                """), {"cid": cid}).fetchall()

            def _build_category(r):
                parts = [p for p in (
                    (r.category_operator or "").strip(),
                    (r.category_mode     or "").strip(),
                    (r.category_power    or "").strip(),
                ) if p]
                return " / ".join(parts) if parts else "Unknown"

            self._ldr_data = [
                (r.callsign,
                 _build_category(r),
                 r.final_score    or 0,
                 r.total_qsos     or 0,
                 r.valid_qsos     or 0,
                 r.duplicate_qsos or 0,
                 r.total_points   or 0,
                 r.multipliers    or 0)
                for r in rows
            ]

            # Populate category filter combobox
            cats = sorted({r[1] for r in self._ldr_data})
            self._ldr_cat_combo["values"] = ["All"] + cats
            if self._ldr_cat_filter.get() not in (["All"] + cats):
                self._ldr_cat_filter.set("All")

            self._apply_leaderboard_filter()
            self._log(f"Leaderboard loaded — {len(self._ldr_data)} entries.", "ok")
        except Exception as e:
            self._log(f"Leaderboard error: {e}", "error")

    def _apply_leaderboard_filter(self):
        flt      = self._ldr_filter.get().strip().upper()
        cat_flt  = self._ldr_cat_filter.get()
        area_flt = self._ldr_area_filter.get()

        for row in self._ldr_tree.get_children():
            self._ldr_tree.delete(row)

        data = self._ldr_data
        if flt:
            data = [r for r in data if flt in r[0].upper()]
        if cat_flt and cat_flt != "All":
            data = [r for r in data if r[1] == cat_flt]
        if area_flt and area_flt != "World":
            if area_flt == "South America":
                # South America but NOT Argentina
                data = [r for r in data
                        if _callsign_area(r[0]) == "South America"]
            else:
                data = [r for r in data
                        if _callsign_area(r[0]) == area_flt]

        # sort — data tuple: (callsign, category, final_score, total_qsos,
        #                     valid_qsos, dupes, points, multipliers)
        col_idx = {"#": -1, "Callsign": 0, "Category": 1,
                   "Final Score": 2, "Total QSOs": 3, "Valid QSOs": 4,
                   "Dupes": 5, "Points": 6, "Multipliers": 7}
        ci = col_idx.get(self._ldr_sort_col, 1)
        if ci >= 0:
            data = sorted(data, key=lambda r: r[ci], reverse=self._ldr_sort_rev)

        # When sorted by category, compute within-category rank
        by_cat = self._ldr_sort_col == "Category"
        cat_rank: dict = {}
        for r in data:
            if by_cat:
                cat_rank[r[1]] = cat_rank.get(r[1], 0) + 1
                rank_label = cat_rank[r[1]]
            else:
                rank_label = None  # filled in below

        overall = 0
        cat_counters: dict = {}
        for r in data:
            overall += 1
            cat_counters[r[1]] = cat_counters.get(r[1], 0) + 1
            rank_label = cat_counters[r[1]] if by_cat else overall
            self._ldr_tree.insert("", "end", values=(
                rank_label, r[0], r[1],
                f"{r[2]:,}", f"{r[3]:,}", f"{r[4]:,}",
                f"{r[5]:,}", f"{r[6]:,}", f"{r[7]:,}"
            ))

    def _sort_leaderboard(self, col: str):
        if self._ldr_sort_col == col:
            self._ldr_sort_rev = not self._ldr_sort_rev
        else:
            self._ldr_sort_col = col
            # text columns sort ascending by default; numeric columns descending
            self._ldr_sort_rev = col not in ("Callsign", "Category")
        self._apply_leaderboard_filter()

    # ─────────────────────────────────────────────────────────────────
    #  Utilities
    # ─────────────────────────────────────────────────────────────────

    def _export_excel(self):
        cid = self._selected_contest_id.get()
        if not cid:
            messagebox.showinfo("No contest", "Select an active contest first.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Export Scores to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="tablaparticipantes.xlsx",
            initialdir="."
        )
        if not out_path:
            return

        db_path = self._db_path.get()

        def _run():
            self._set_status("Exporting to Excel…", busy=True)
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from collections import defaultdict
                from src.database.db_manager import DatabaseManager
                from sqlalchemy import text as sql_text

                db = DatabaseManager(db_path)
                with db.get_session() as session:
                    rows = session.execute(sql_text("""
                        SELECT l.callsign, l.name, l.category_operator,
                               l.category_power, l.category_mode, l.operators,
                               l.created_by, l.club, l.extra_data, l.file_path,
                               l.address_country, l.email,
                               COALESCE(s.final_score, 0) AS final_score
                        FROM logs l
                        LEFT JOIN scores s ON s.log_id = l.id
                        WHERE l.contest_id = :cid
                        ORDER BY final_score DESC
                    """), {"cid": cid}).fetchall()

                    # Continent and country lookup via CTY/DXCC data
                    continent_cache = {}
                    dxcc_country_cache = {}
                    try:
                        from src.services.callsign_lookup import CallsignLookupService
                        lookup_svc = CallsignLookupService(session)
                        for r in rows:
                            cs = r.callsign or ""
                            if cs and cs not in continent_cache:
                                info = lookup_svc.lookup_callsign(cs)
                                continent_cache[cs] = info.get("continent", "") if info else ""
                                dxcc_country_cache[cs] = info.get("country_name", "") if info else ""
                    except Exception:
                        pass

                def _get_soapbox(extra_data):
                    if isinstance(extra_data, dict):
                        return (extra_data.get("soapbox") or
                                extra_data.get("SOAPBOX") or None)
                    return None

                # Build enriched data list (sorted by score desc already)
                data = []
                for r in rows:
                    cs = r.callsign or ""
                    is_checklog = (r.category_operator or "").upper() == "CHECKLOG"
                    cat_key = (
                        (r.category_operator or "").upper(),
                        (r.category_power or "").upper(),
                        (r.category_mode or "").upper(),
                    )
                    data.append({
                        "callsign":    cs,
                        "name":        r.name or "",
                        "categoria":   r.category_operator or "",
                        "subcategoria":r.category_power or "",
                        "modo1":       r.category_mode or "",
                        "opes":        r.operators or "",
                        "soft":        r.created_by or "",
                        "clubes":      r.club or "",
                        "soapbox":     _get_soapbox(r.extra_data),
                        "archivo":     Path(r.file_path).name if r.file_path else "",
                        "miPais":      dxcc_country_cache.get(cs, "") or r.address_country or "",
                        "miCont":      continent_cache.get(cs, ""),
                        "GranTotal":   r.final_score,
                        "email":       r.email or "",
                        "area_name":   _callsign_area(cs),
                        "cat_key":     cat_key,
                        "is_checklog": is_checklog,
                        "PosPais":     0,
                        "PosContinente": 0,
                        "area":        0,
                        "PosMundo":    0,
                    })

                # Compute per-category rankings (CHECKLOG stays 0)
                cat_groups = defaultdict(list)
                for idx, d in enumerate(data):
                    if not d["is_checklog"]:
                        cat_groups[d["cat_key"]].append(idx)

                for cat_key, indices in cat_groups.items():
                    sorted_idx = sorted(indices,
                                        key=lambda i: data[i]["GranTotal"],
                                        reverse=True)
                    world_rank    = 0
                    country_rank  = defaultdict(int)
                    cont_rank     = defaultdict(int)
                    area_rank     = defaultdict(int)
                    for idx in sorted_idx:
                        d = data[idx]
                        world_rank += 1
                        d["PosMundo"]      = world_rank
                        country_rank[d["miPais"]] += 1
                        d["PosPais"]       = country_rank[d["miPais"]]
                        cont_rank[d["miCont"]] += 1
                        d["PosContinente"] = cont_rank[d["miCont"]]
                        area_rank[d["area_name"]] += 1
                        d["area"]          = area_rank[d["area_name"]]

                # Write workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "tablaparticipantes"

                col_headers = [
                    "Id", "Estacion", "Nombre", "categoria", "subcategoria",
                    "modo1", "opes", "Soft", "clubes", "soapbox", "Archivo",
                    "miPais", "miCont", "GranTotal", "email",
                    "PosPais", "PosContinente", "area", "PosMundo",
                ]
                hdr_font = Font(bold=True, color="FFFFFF")
                hdr_fill = PatternFill("solid", fgColor="1A3A5C")
                for col_num, h in enumerate(col_headers, 1):
                    cell = ws.cell(1, col_num, h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal="center")

                for i, d in enumerate(data, 1):
                    ws.append([
                        i, d["callsign"], d["name"], d["categoria"],
                        d["subcategoria"], d["modo1"], d["opes"], d["soft"],
                        d["clubes"], d["soapbox"], d["archivo"],
                        d["miPais"], d["miCont"], d["GranTotal"], d["email"],
                        d["PosPais"], d["PosContinente"], d["area"], d["PosMundo"],
                    ])

                # Auto-size columns
                for col in ws.columns:
                    max_len = max(
                        (len(str(cell.value or "")) for cell in col), default=8
                    )
                    ws.column_dimensions[
                        col[0].column_letter
                    ].width = min(max_len + 2, 45)

                wb.save(out_path)
                self._log(f"Exported {len(data)} rows → {out_path}", "ok")

            except ImportError:
                self._log(
                    "openpyxl is not installed. Run: pip install openpyxl", "error")
                messagebox.showerror(
                    "Missing dependency",
                    "openpyxl is required for Excel export.\n\nRun: pip install openpyxl")
            except Exception as e:
                import traceback
                self._log(f"Export error: {e}", "error")
                self._log(traceback.format_exc(), "error")
            finally:
                self._set_status("Ready")

        threading.Thread(target=_run, daemon=True).start()

    def _export_qso_excel(self):
        """Export a per-participant QSO Excel report with an Observations column."""
        cid = self._selected_contest_id.get()
        if not cid:
            messagebox.showinfo("No contest", "Select an active contest first.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Export QSO Report to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="qso_report.xlsx",
            initialdir="."
        )
        if not out_path:
            return

        db_path = self._db_path.get()

        def _run():
            self._set_status("Building QSO report…", busy=True)
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                from src.database.db_manager import DatabaseManager
                from sqlalchemy import text as sql_text

                # Human-readable labels for each validation status
                STATUS_LABEL = {
                    "valid":              "VALID",
                    "duplicate":          "DUPLICATE",
                    "invalid":            "INVALID",
                    "invalid_callsign":   "INVALID CALLSIGN",
                    "invalid_exchange":   "INVALID EXCHANGE",
                    "out_of_period":      "OUT OF PERIOD",
                    "invalid_band":       "INVALID BAND",
                    "invalid_mode":       "INVALID MODE",
                    "not_in_log":         "NIL (NOT IN LOG)",
                    "time_mismatch":      "TIME MISMATCH",
                    "exchange_mismatch":  "EXCHANGE MISMATCH",
                    "unique_call":        "UNIQUE CALL",
                }

                # Fill colours per status
                STATUS_FILL = {
                    "valid":             PatternFill("solid", fgColor="C6EFCE"),  # light green
                    "duplicate":         PatternFill("solid", fgColor="FFEB9C"),  # yellow
                    "not_in_log":        PatternFill("solid", fgColor="FFC7CE"),  # light red
                    "invalid_callsign":  PatternFill("solid", fgColor="FFC7CE"),
                    "invalid_exchange":  PatternFill("solid", fgColor="FFC7CE"),
                    "out_of_period":     PatternFill("solid", fgColor="FFC7CE"),
                    "invalid_band":      PatternFill("solid", fgColor="FFC7CE"),
                    "invalid_mode":      PatternFill("solid", fgColor="FFC7CE"),
                    "invalid":           PatternFill("solid", fgColor="FFC7CE"),
                    "time_mismatch":     PatternFill("solid", fgColor="FFEB9C"),
                    "exchange_mismatch": PatternFill("solid", fgColor="FFEB9C"),
                    "unique_call":       PatternFill("solid", fgColor="DDEBF7"),  # light blue
                }

                db = DatabaseManager(db_path)
                with db.get_session() as session:
                    # ── Summary data ──────────────────────────────────────────────
                    summary_rows = session.execute(sql_text("""
                        SELECT l.id AS log_id, l.callsign, l.name,
                               l.category_operator, l.category_power,
                               l.category_mode,
                               l.operators, l.club, l.extra_data,
                               l.address_country,
                               COALESCE(s.total_qsos,   0) AS total_qsos,
                               COALESCE(s.valid_qsos,   0) AS valid_qsos,
                               COALESCE(s.duplicate_qsos, 0) AS dupe_qsos,
                               COALESCE(s.not_in_log_qsos, 0) AS nil_qsos,
                               COALESCE(s.invalid_qsos, 0) AS invalid_qsos,
                               COALESCE(s.final_score,  0) AS final_score
                        FROM logs l
                        LEFT JOIN scores s ON s.log_id = l.id
                        WHERE l.contest_id = :cid
                        ORDER BY final_score DESC
                    """), {"cid": cid}).fetchall()

                    # ── Country / continent lookup via CTY data ───────────────────
                    continent_cache = {}
                    country_cache = {}
                    try:
                        from src.services.callsign_lookup import CallsignLookupService
                        lookup_svc = CallsignLookupService(session)
                        for r in summary_rows:
                            cs = r.callsign or ""
                            if cs and cs not in continent_cache:
                                info = lookup_svc.lookup_callsign(cs)
                                continent_cache[cs] = info.get("continent", "") if info else ""
                                country_cache[cs]   = info.get("country_name", "") if info else ""
                    except Exception:
                        pass

                    # ── Contact data ──────────────────────────────────────────────
                    contact_rows = session.execute(sql_text("""
                        SELECT c.log_id, c.qso_date, c.qso_time, c.band, c.mode,
                               c.call_sent, c.rst_sent, c.exchange_sent,
                               c.call_received, c.rst_received, c.exchange_received,
                               COALESCE(c.points, 0) AS points,
                               c.validation_status, c.validation_notes
                        FROM contacts c
                        JOIN logs l ON c.log_id = l.id
                        WHERE l.contest_id = :cid
                        ORDER BY c.log_id, c.qso_date, c.qso_time
                    """), {"cid": cid}).fetchall()

                # Group contacts by log_id
                from collections import defaultdict
                contacts_by_log = defaultdict(list)
                for row in contact_rows:
                    contacts_by_log[row.log_id].append(row)

                def _get_soapbox_qso(extra_data):
                    if isinstance(extra_data, dict):
                        val = extra_data.get("soapbox") or extra_data.get("SOAPBOX")
                        if isinstance(val, list):
                            return " | ".join(val)
                        return val or ""
                    return ""

                # ── Build workbook ────────────────────────────────────────────────
                wb = openpyxl.Workbook()

                # ── Sheet 1: Summary ──────────────────────────────────────────────
                ws_sum = wb.active
                ws_sum.title = "Summary"

                hdr_font  = Font(bold=True, color="FFFFFF")
                hdr_fill  = PatternFill("solid", fgColor="1A3A5C")
                ctr_align = Alignment(horizontal="center", vertical="center")

                sum_headers = [
                    "#", "Callsign", "Name", "Category", "Power", "Mode",
                    "Operators", "Club", "Soapbox", "Country", "Continent",
                    "Total QSOs", "Valid QSOs", "Duplicates", "NIL", "Invalid",
                    "Final Score",
                ]
                for col_num, h in enumerate(sum_headers, 1):
                    cell = ws_sum.cell(1, col_num, h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = ctr_align

                for idx, r in enumerate(summary_rows, 1):
                    is_checklog = (r.category_operator or "").upper() == "CHECKLOG"
                    cs = r.callsign or ""
                    ws_sum.append([
                        idx,
                        cs,
                        r.name or "",
                        r.category_operator or "",
                        r.category_power or "",
                        r.category_mode or "",
                        r.operators or "",
                        r.club or "",
                        _get_soapbox_qso(r.extra_data),
                        country_cache.get(cs, "") or r.address_country or "",
                        continent_cache.get(cs, ""),
                        r.total_qsos,
                        r.valid_qsos,
                        r.dupe_qsos,
                        r.nil_qsos,
                        r.invalid_qsos,
                        0 if is_checklog else r.final_score,
                    ])

                # Auto-size summary columns
                for col in ws_sum.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws_sum.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

                # Freeze header row
                ws_sum.freeze_panes = "A2"

                # ── Per-participant sheets ────────────────────────────────────────
                qso_headers = [
                    "#", "Date", "Time", "Band", "Mode",
                    "Call Sent", "RST Sent", "Exch Sent",
                    "Call Rcvd", "RST Rcvd", "Exch Rcvd",
                    "Points", "Observations",
                ]

                used_titles = set()

                def _safe_sheet_name(callsign: str) -> str:
                    """Return a valid, unique Excel sheet name (max 31 chars)."""
                    # Strip characters not allowed in sheet names
                    for ch in "/\\?*[]:'":
                        callsign = callsign.replace(ch, "_")
                    title = callsign[:31]
                    base, counter = title, 2
                    while title in used_titles:
                        suffix = f"_{counter}"
                        title = base[:31 - len(suffix)] + suffix
                        counter += 1
                    used_titles.add(title)
                    return title

                for r in sorted(summary_rows, key=lambda r: (r.callsign or "").upper()):
                    log_id   = r.log_id
                    callsign = r.callsign or f"LOG{log_id}"
                    contacts = contacts_by_log.get(log_id, [])

                    ws = wb.create_sheet(title=_safe_sheet_name(callsign))

                    # Sheet title row
                    ws.merge_cells(start_row=1, start_column=1,
                                   end_row=1,   end_column=len(qso_headers))
                    title_cell = ws.cell(1, 1, f"{callsign}  —  QSO Report")
                    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
                    title_cell.fill = PatternFill("solid", fgColor="1A3A5C")
                    title_cell.alignment = Alignment(horizontal="left", vertical="center")
                    ws.row_dimensions[1].height = 22

                    # Column headers (row 2)
                    for col_num, h in enumerate(qso_headers, 1):
                        cell = ws.cell(2, col_num, h)
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                        cell.alignment = ctr_align

                    # QSO rows (starting at row 3)
                    for qso_num, c in enumerate(contacts, 1):
                        status   = (c.validation_status or "valid").lower()
                        label    = STATUS_LABEL.get(status, status.upper())
                        row_fill = STATUS_FILL.get(status)
                        row_num  = qso_num + 2

                        values = [
                            qso_num,
                            c.qso_date or "",
                            c.qso_time or "",
                            c.band or "",
                            c.mode or "",
                            c.call_sent or "",
                            c.rst_sent or "",
                            c.exchange_sent or "",
                            c.call_received or "",
                            c.rst_received or "",
                            c.exchange_received or "",
                            c.points,
                            label,
                        ]
                        ws.append(values)

                        if row_fill:
                            for col_num in range(1, len(qso_headers) + 1):
                                ws.cell(row_num, col_num).fill = row_fill

                    # Auto-size columns
                    col_widths = [4, 12, 6, 6, 5, 14, 6, 10, 14, 6, 10, 7, 20]
                    for col_num, width in enumerate(col_widths, 1):
                        ws.column_dimensions[
                            get_column_letter(col_num)
                        ].width = width

                    # Freeze title + header rows
                    ws.freeze_panes = "A3"

                wb.save(out_path)
                self._log(
                    f"QSO report exported: {len(summary_rows)} participants, "
                    f"{len(contact_rows)} QSOs → {out_path}", "ok"
                )
                messagebox.showinfo(
                    "Export complete",
                    f"QSO report saved to:\n{out_path}\n\n"
                    f"{len(summary_rows)} participants / {len(contact_rows)} QSOs"
                )

            except ImportError:
                self._log("openpyxl is not installed. Run: pip install openpyxl", "error")
                messagebox.showerror(
                    "Missing dependency",
                    "openpyxl is required for Excel export.\n\nRun: pip install openpyxl")
            except Exception as e:
                import traceback
                self._log(f"QSO report export error: {e}", "error")
                self._log(traceback.format_exc(), "error")
                messagebox.showerror("Export error", str(e))
            finally:
                self._set_status("Ready")

        threading.Thread(target=_run, daemon=True).start()

    def _export_scores_csv(self):
        cid = self._selected_contest_id.get()
        if not cid:
            messagebox.showinfo("No contest", "Select an active contest first.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Export Scores to CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="scores_export.csv",
            initialdir="."
        )
        if not out_path:
            return

        db_path = self._db_path.get()

        def _run():
            self._set_status("Exporting scores to CSV…", busy=True)
            try:
                import csv
                from src.database.db_manager import DatabaseManager
                from sqlalchemy import text as sql_text

                db = DatabaseManager(db_path)
                with db.get_session() as session:
                    result = session.execute(sql_text("""
                        SELECT scores.*, logs.*
                        FROM scores
                        JOIN logs ON scores.log_id = logs.id
                        WHERE logs.contest_id = :cid
                    """), {"cid": cid})
                    rows = result.fetchall()
                    columns = list(result.keys())

                with open(out_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(columns)
                    writer.writerows(rows)

                self._log(f"Exported {len(rows)} rows → {out_path}", "ok")
                messagebox.showinfo("Export complete",
                                    f"Exported {len(rows)} rows to:\n{out_path}")
            except Exception as e:
                import traceback
                self._log(f"CSV export error: {e}", "error")
                self._log(traceback.format_exc(), "error")
                messagebox.showerror("Export error", str(e))
            finally:
                self._set_status("Ready")

        threading.Thread(target=_run, daemon=True).start()

    def _select_db(self):
        path = filedialog.askopenfilename(
            title="Select database file",
            filetypes=[("SQLite", "*.db"), ("All", "*.*")],
            initialdir="."
        )
        if path:
            self._db_path.set(path)
            self._log(f"Database changed to: {path}", "ok")
            self._init_database()
            self._refresh_contests()


# ─────────────────────────────────────────────────────────────────────

def main():
    # Change working directory to script location so relative paths work
    os.chdir(Path(__file__).parent)
    app = SA10App()
    app.mainloop()


if __name__ == "__main__":
    main()
